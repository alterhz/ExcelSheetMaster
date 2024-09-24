import openpyxl


class ExcelSheetHandler:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def create_sheet(self):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            if self.sheet_name not in workbook.sheetnames:
                new_sheet = workbook.create_sheet(title=self.sheet_name)
                print("页签不存在，创建")
            return workbook
        except FileNotFoundError:
            print(f"无法找到指定的工作簿：{self.file_name}")
            return None

    def insert_column_header(self, index, cs, type_name, name, note, column_width=0):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            sheet.insert_cols(index, 1)
            sheet.cell(row=1, column=index).value = cs
            sheet.cell(row=2, column=index).value = type_name
            sheet.cell(row=3, column=index).value = name
            sheet.cell(row=4, column=index).value = note
            if column_width > 0:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = column_width
            print(f"插入列。index:{index}, name:{name}")

    def write_column_header(self, index, cs, type_name, name, note, column_width=0):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            sheet.cell(row=1, column=index).value = cs
            sheet.cell(row=2, column=index).value = type_name
            sheet.cell(row=3, column=index).value = name
            sheet.cell(row=4, column=index).value = note
            if column_width > 0:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = column_width
            print(f"写入列。index:{index}, name:{name}")

    def set_header_color(self, color):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            for row in range(1, 5):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(start_color=color,
                                                                                       end_color=color,
                                                                                       fill_type="solid")

    def delete_column(self, index):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            if 1 <= index <= sheet.max_column:
                sheet.delete_cols(index)
            else:
                print("指定的列索引无效")

    def move_column(self, from_index, to_index):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            if 1 <= from_index <= sheet.max_column and 1 <= to_index <= sheet.max_column:
                column_to_move = [row[from_index - 1].value for row in sheet.iter_rows(values_only=True)]
                sheet.delete_cols(from_index)
                sheet.insert_cols(to_index)
                for row_num, value in enumerate(column_to_move, start=1):
                    sheet.cell(row=row_num, column=to_index).value = value
            else:
                print("指定的列索引无效")

    def get_sheet_header(self):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            data_collection = []
            column_count = sheet.max_column
            for i in range(1, column_count + 1):
                column_data = {
                    "columnIndex": i,
                    "exportType": sheet.cell(row=1, column=i).value,
                    "dataType": sheet.cell(row=2, column=i).value,
                    "name": sheet.cell(row=3, column=i).value,
                    "note": sheet.cell(row=4, column=i).value
                }
                data_collection.append(column_data)
            return data_collection
        else:
            return None

    def get_column_count(self):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            return sheet.max_column
        else:
            return None

    def clear_data_rows(self):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            if last_row >= 5:
                for row in range(5, last_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).value = None

    def get_first_row_data_by_column_values(self, *find_values):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            result_dict = {}
            for i in range(1, last_row + 1):
                all_columns_match = True
                j = 0
                while j < len(find_values):
                    column_name = find_values[j]
                    value_to_find = find_values[j + 1]
                    column_index = self.get_column_index_by_name(column_name)
                    if column_index == 0:
                        print(f"找不到指定的列名：{column_name}")
                        return {}
                    cell_value = sheet.cell(row=i, column=column_index).value
                    if cell_value != value_to_find:
                        all_columns_match = False
                        break
                    j += 2
                if all_columns_match:
                    column_count = sheet.max_column
                    for column_index in range(1, column_count + 1):
                        key = sheet.cell(row=3, column=column_index).value
                        value = sheet.cell(row=i, column=column_index).value
                        result_dict[key] = value
                    result_dict["row"] = i
                    break
            return result_dict
        else:
            return {}

    def get_last_row_data_by_column_values(self, *find_values):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            result_dict = {}
            found = False
            if len(find_values) == 0:
                found_row = last_row
                found = True
            else:
                for i in range(last_row, 0, -1):
                    all_columns_match = True
                    j = 0
                    while j < len(find_values):
                        column_name = find_values[j]
                        value_to_find = find_values[j + 1]
                        column_index = self.get_column_index_by_name(column_name)
                        if column_index == 0:
                            print(f"找不到指定的列名：{column_name}")
                            return {}
                        cell_value = sheet.cell(row=i, column=column_index).value
                        if cell_value != value_to_find:
                            all_columns_match = False
                            break
                        j += 2
                    if all_columns_match:
                        found_row = i
                        found = True
                        break
            if found:
                column_count = sheet.max_column
                for column_index in range(1, column_count + 1):
                    key = sheet.cell(row=3, column=column_index).value
                    value = sheet.cell(row=found_row, column=column_index).value
                    result_dict[key] = value
                result_dict["row"] = found_row
            return result_dict
        else:
            return {}

    def get_all_data(self):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            result_collection = []
            for i in range(5, last_row + 1):
                result_dict = {}
                column_count = sheet.max_column
                for column_index in range(1, column_count + 1):
                    key = sheet.cell(row=3, column=column_index).value
                    value = sheet.cell(row=i, column=column_index).value
                    result_dict[key] = value
                result_dict["row"] = i
                result_collection.append(result_dict)
            return result_collection
        else:
            return None

    def get_all_rows_data_by_column_values(self, *find_values):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            result_collection = []
            found = False
            if len(find_values) == 0:
                found = True
            else:
                result = find_values
                for i in range(5, last_row + 1):
                    all_columns_match = True
                    j = 0
                    while j < len(result):
                        column_name = result[j]
                        value_to_find = result[j + 1]
                        column_index = self.get_column_index_by_name(column_name)
                        if column_index == 0:
                            print(f"找不到指定的列名：{column_name}")
                            return []
                        cell_value = sheet.cell(row=i, column=column_index).value
                        if str(cell_value) != str(value_to_find):
                            all_columns_match = False
                            break
                        j += 2
                    if all_columns_match:
                        found = True
                        result_dict = {}
                        column_count = sheet.max_column
                        for column_index in range(1, column_count + 1):
                            key = sheet.cell(row=3, column=column_index).value
                            value = sheet.cell(row=i, column=column_index).value
                            result_dict[key] = value
                        result_dict["row"] = i
                        result_collection.append(result_dict)
            if found:
                return result_collection
            else:
                return None
        else:
            return None

    def get_column_values_by_column_names(self, *column_names):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            temp_col = []
            for i in range(5, last_row + 1):
                temp_dict = {"row": i}
                for column_name in column_names:
                    column_index = self.get_column_index_by_name(column_name)
                    if column_index == 0:
                        print(f"找不到指定的列名：{column_name}")
                        return None
                    temp_dict[column_name] = sheet.cell(row=i, column=column_index).value
                temp_col.append(temp_dict)
            return temp_col
        else:
            return None

    def get_unique_column_values(self, column_name):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            unique_col = []
            column_index = self.get_column_index_by_name(column_name)
            if column_index == 0:
                print(f"找不到指定的列名：{column_name}")
                return None
            for i in range(5, last_row + 1):
                temp_value = sheet.cell(row=i, column=column_index).value
                if temp_value not in unique_col:
                    unique_col.append(temp_value)
            return unique_col
        else:
            return None

    def is_in_collection(self, value, col):
        for item in col:
            if item == value:
                return True
        return False

    def get_max_value_from_column(self, column_name):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            column_index = self.get_column_index_by_name(column_name)
            if column_index == 0:
                print(f"找不到指定的列名：{column_name}")
                return -99999999
            values = []
            for i in range(5, sheet.max_row + 1):
                value = sheet.cell(row=i, column=column_index).value
                if value is not None:
                    values.append(int(value))
            max_value = -99999999
            for value in values:
                if value > max_value:
                    max_value = value
            return max_value
        else:
            return -99999999

    def alloc_sn(self):
        max_sn = self.get_max_value_from_column("sn")
        return max_sn + 1

    def get_max_sn(self):
        max_sn = self.get_max_value_from_column("sn")
        return max_sn

    def insert_row_and_write_dict_data(self, row_number, dict_data, insert=True):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            if row_number == -1:
                row_number = sheet.max_row + 1
            elif row_number < 5:
                print("插入行号不能小于 5。")
                return -1
            if insert:
                sheet.insert_rows(row_number)
            for key, value in dict_data.items():
                column_name = key
                column_index = self.get_column_index_by_name(column_name)
                if column_index != 0:
                    target_cell = sheet.cell(row=row_number, column=column_index)
                    target_cell.number_format = "@"
                    target_cell.value = value
            return row_number
        else:
            return -1

    def insert_row_and_write_param_data(self, row_number, *column_value_pairs):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            if row_number == -1:
                row_number = sheet.max_row + 1
            sheet.insert_rows(row_number)
            i = 0
            while i < len(column_value_pairs):
                column_name = column_value_pairs[i]
                value = column_value_pairs[i + 1]
                column_index = self.get_column_index_by_name(column_name)
                if column_index != 0:
                    sheet.cell(row=row_number, column=column_index).value = value
                i += 2

    def get_column_index_by_name(self, column_name):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_column = sheet.max_column
            for i in range(1, last_column + 1):
                if sheet.cell(row=3, column=i).value == column_name:
                    return i
            return 0
        else:
            return 0

    def move_to_row(self, row_number):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            last_row = sheet.max_row
            if row_number < 1:
                row_number = 1
            elif row_number > last_row:
                row_number = last_row
            sheet.cell(row=row_number, column=1).select()

    def set_row_color(self, row_number, color):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            for col_num in range(1, sheet.max_column + 1):
                sheet.cell(row=row_number, column=col_num).fill = openpyxl.styles.PatternFill(start_color=color,
                                                                                              end_color=color,
                                                                                              fill_type="solid")

    def move_to_row_and_set_color(self, row_number, color):
        self.move_to_row(row_number)
        self.set_row_color(row_number, color)

    def get_max_row_number(self):
        workbook = self.create_sheet()
        if workbook:
            sheet = workbook[self.sheet_name]
            return sheet.max_row
        else:
            return None


if __name__ == '__main__':
    handler = ExcelSheetHandler("your_workbook_name.xlsx", "your_sheet_name")
# 调用各种方法进行测试
