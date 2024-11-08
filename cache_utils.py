import logging
import os
import time
from multiprocessing import Queue, Process

import openpyxl

import excel_utils
import os_utils
from excel_sheet_handler import ExcelSheetHandler
from excel_utils import get_excel_sheet_names

CACHE_EXCEL_NAME = "cache.xlsx"
CONFIG_SHEET_NAME = "config"
ALL_PATH_SHEET_NAME = "allPath"

g_cache_config: ExcelSheetHandler = None
g_cache_all_path: ExcelSheetHandler = None
g_cache_all_sheet: dict = {}

qIn = Queue()
qOut = Queue()
process = None
waiting_run_excels = []


def close_cache():
    global g_cache_config, g_cache_all_path
    g_cache_config = None
    g_cache_all_path = None
    g_cache_all_sheet.clear()
    ExcelSheetHandler.close_all_workbook()


def create_excel_sheet(excel_name, sheet_name=None):
    """
    检查缓存文件是否存在，如果不存在则创建，并可根据需求创建指定名称的工作表。

    参数：
    excel_name：Excel 文件名称。
    sheet_name（可选）：要创建的工作表名称，如果不提供则不进行创建工作表的操作。

    返回：
    bool：如果创建了文件或者页签返回 True，否则返回 False。
    """
    sheet_created = False
    if not os.path.exists(excel_name):
        wb = openpyxl.Workbook()
        wb.save(excel_name)
    workbook = openpyxl.load_workbook(excel_name)
    if sheet_name and sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        sheet_created = True
        workbook.save(excel_name)
    return sheet_created


def get_config_cache_sheet():
    global g_cache_config
    if g_cache_config is None:
        create_sheet = create_excel_sheet(CACHE_EXCEL_NAME, CONFIG_SHEET_NAME)
        g_cache_config = ExcelSheetHandler(CACHE_EXCEL_NAME, CONFIG_SHEET_NAME)
        if create_sheet:
            g_cache_config.insert_column_header(1, "cs", "String", "key", "键", 20)
            g_cache_config.insert_column_header(2, "cs", "String", "value", "值", 30)
            g_cache_config.save_workbook()
            logging.info(f"创建缓存文件 {CACHE_EXCEL_NAME} 的工作表 {CONFIG_SHEET_NAME}，并插入列头。")
    return g_cache_config


def get_all_path_cache_sheet():
    global g_cache_all_path
    if g_cache_all_path is None:
        create_sheet = create_excel_sheet(CACHE_EXCEL_NAME, ALL_PATH_SHEET_NAME)
        g_cache_all_path = ExcelSheetHandler(CACHE_EXCEL_NAME, ALL_PATH_SHEET_NAME)
        if create_sheet:
            g_cache_all_path.insert_column_header(1, "cs", "String", "path", "路径", 60)
            g_cache_all_path.insert_column_header(2, "cs", "String", "sheet_name", "页签名称,自动生成", 20)
            g_cache_all_path.insert_column_header(3, "cs", "Boolean", "includeSubDir", "是否包含子文件夹", 20)
            g_cache_all_path.insert_column_header(4, "cs", "String", "desc", "描述", 60)
            g_cache_all_path.save_workbook()
            logging.info(f"创建缓存文件 {CACHE_EXCEL_NAME} 的工作表 {ALL_PATH_SHEET_NAME}，并插入列头。")
    return g_cache_all_path


def get_cache_sheet(sheet_name):
    global g_cache_all_sheet
    if g_cache_all_sheet.get(sheet_name) is None:
        create_sheet = create_excel_sheet(CACHE_EXCEL_NAME, sheet_name)
        g_cache_all_sheet[sheet_name] = ExcelSheetHandler(CACHE_EXCEL_NAME, sheet_name)
        if create_sheet:
            g_cache_all_sheet[sheet_name].insert_column_header(1, "cs", "String", "name", "文件名", 50)
            g_cache_all_sheet[sheet_name].insert_column_header(2, "cs", "String", "lastModified", "文件修改时间", 50)
            g_cache_all_sheet[sheet_name].insert_column_header(3, "cs", "String", "sheets", "页签列表", 50)
            g_cache_all_sheet[sheet_name].insert_column_header(4, "cs", "Boolean", "need_update", "需要更新页签列表",
                                                               20)
            g_cache_all_sheet[sheet_name].save_workbook()
            logging.info(f"创建缓存文件 {CACHE_EXCEL_NAME} 的工作表 {sheet_name}，并插入列头。")
    return g_cache_all_sheet[sheet_name]


def remove_cache_sheet(sheet_name):
    if sheet_name in g_cache_all_sheet:
        g_cache_all_sheet[sheet_name].remove_sheet()
        del g_cache_all_sheet[sheet_name]
        logging.info(f"删除缓存文件 {CACHE_EXCEL_NAME} 的工作表 {sheet_name}。")


def set_config_value(key, value):
    config_cache = get_config_cache_sheet()
    data = config_cache.get_all_data()
    for row in data:
        if row["key"] == key:
            row["value"] = value
            config_cache.write_row_data(row["row"], row)
            config_cache.save_workbook()
            return
    row = {"key": key, "value": value}
    config_cache.write_row_data(g_cache_config.get_max_row_number() + 1, row)
    config_cache.save_workbook()


def get_config_value(key):
    config_cache = get_config_cache_sheet()
    data = config_cache.get_all_data()
    for row in data:
        if row["key"] == key:
            return row["value"]
    return None


def set_path_data(path, sheet_name, include_subdir, desc):
    all_path_sheet = get_all_path_cache_sheet()
    data = all_path_sheet.get_all_data()
    for row in data:
        if row["path"] == path:
            row["includeSubDir"] = include_subdir
            row["desc"] = desc
            all_path_sheet.write_row_data(row["row"], row)
            all_path_sheet.save_workbook()
            return

    row = {"path": path, "includeSubDir": include_subdir, "desc": desc, "sheet_name": sheet_name}
    all_path_sheet.write_row_data(all_path_sheet.get_max_row_number() + 1, row)
    all_path_sheet.save_workbook()


def del_path_data(path: str):
    """删除指定路径的缓存数据"""
    all_path_sheet = get_all_path_cache_sheet()
    data = all_path_sheet.get_all_data()
    for row in data:
        if row["path"] == path:
            all_path_sheet.delete_rows([row["row"]])
            all_path_sheet.save_workbook()
            return


def get_path_data(path):
    all_path_sheet = get_all_path_cache_sheet()
    data = all_path_sheet.get_all_data()
    for row in data:
        if row["path"] == path:
            return row
    return None


def get_all_path_data():
    all_path_sheet = get_all_path_cache_sheet()
    return all_path_sheet.get_all_data()


def get_first_path():
    all_path_sheet = get_all_path_cache_sheet()
    data = all_path_sheet.get_all_data()
    if len(data) > 0:
        return data[0]["path"]
    return None


def get_path_sheet_name(path):
    all_path_sheet = get_all_path_cache_sheet()
    data = all_path_sheet.get_all_data()
    for row in data:
        if row["path"] == path:
            return row["sheet_name"]
    return None


def compute_cache_data():
    # 计算缓存数据
    use_path = get_config_value("usePath")
    use_sheet_name = get_path_sheet_name(use_path)
    sheet_handler = get_cache_sheet(use_sheet_name)
    data = sheet_handler.get_all_data()
    # 遍历data转为字典key为文件名,value为行数据
    excel_map = {}
    for row in data:
        excel_map[row["name"]] = row

    modified_count = 0
    names = os_utils.get_current_file_names(use_path, ".xlsx")
    for fullname in names:
        name = os_utils.get_filename_from_path(fullname)
        last_modified = os.path.getmtime(fullname)
        # 转为时间格式
        last_modified = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(last_modified))
        # data中获取对应的数据
        if name in excel_map:
            row = excel_map[name]
            if row is None:
                continue
            if row["lastModified"] == last_modified:
                continue
            else:
                # 获取页签列表，更新页签列表和修改时间
                row["lastModified"] = last_modified
                row["need_update"] = True
                sheet_handler.write_row_data(row["row"], row)
                modified_count += 1
                logging.info("更新处理文件：" + name)
        else:
            row = {"cs": "cs", "name": name, "lastModified": last_modified, "sheets": "", "need_update": True}
            sheet_handler.write_row_data(sheet_handler.get_max_row_number() + 1, row)
            modified_count += 1
            logging.info("新增待处理文件：" + name)

    # 转为names未only_names，去掉路径，只要文件名称
    only_names = []
    for name in names:
        only_names.append(os_utils.get_filename_from_path(name))

    # 不存在的文件，要删除数据。记录行号到数组中
    delete_rows = []
    for row in data:
        if row["name"] not in only_names:
            delete_rows.append(row["row"])
            logging.info("删除待处理文件：" + row["name"])

    # 删除数据
    sheet_handler.delete_rows(delete_rows)

    # 保存
    sheet_handler.save_workbook()

    # 构建要处理的文件列表
    use_path = get_config_value("usePath")
    use_sheet_name = get_path_sheet_name(use_path)
    # 获取待处理的所有文件列表
    data = sheet_handler.get_all_data()
    # 遍历data转为字典key为文件名,value为行数据
    global waiting_run_excels
    waiting_run_excels = []
    for row in data:
        if row["need_update"]:
            waiting_run_excels.append(use_path + "/" + row["name"])


def is_all_empty():
    global waiting_run_excels
    return qIn.empty() and qOut.empty() and waiting_run_excels.__len__() == 0


def get_waiting_run_excel_count():
    return waiting_run_excels.__len__()


def run_thread():
    if qIn.empty():
        # 获取10条数据waiting_run_excels
        global waiting_run_excels
        if waiting_run_excels.__len__() > 0:
            if len(waiting_run_excels) >= 10:
                qIn.put(waiting_run_excels[:10])
                del waiting_run_excels[:10]
                logging.debug(f"请求线程处理一批工作簿，数量：{10}，剩余：{waiting_run_excels.__len__()}")
            else:
                qIn.put(waiting_run_excels)
                logging.debug(f"请求线程处理最后一批工作簿，数量：{waiting_run_excels.__len__()}")
                waiting_run_excels = []

    if not qOut.empty():
        t1 = time.time()
        excel_sheets = qOut.get()
        use_path = get_config_value("usePath")
        use_path_name = get_path_sheet_name(use_path)
        sheet_handler = get_cache_sheet(use_path_name)
        data = sheet_handler.get_all_data()
        count = 0
        # 打印
        for excel_name, sheet_names in excel_sheets.items():
            # 获取路径，不包含文件名
            path = os.path.dirname(excel_name)
            if use_path == path:
                # logging.info(f"{path}文件：{excel_name}，页签数：{sheet_names}")
                for row in data:
                    if row["name"] == os_utils.get_filename_from_path(excel_name):
                        row["sheets"] = filter_sheet_names(sheet_names)
                        row["need_update"] = False
                        sheet_handler.write_row_data(row["row"], row)
                        count += 1
                        # logging.info(f"更新页签列表：{excel_name}，页签数：{sheet_names}")
                        break
            else:
                logging.warning(f"路径不匹配：{path}，配置路径：{use_path}")
        # 保存
        if count > 0:
            sheet_handler.save_workbook()
        t2 = time.time()
        logging.info(f"处理缓存数据耗时：{t2 - t1:.2}秒")


def start_back_thread():
    global process
    process = Process(target=worker, args=(qIn, qOut))
    process.start()


def stop_back_thread():
    global process, qIn
    qIn.put(None)
    if process:
        process.join()
        logging.info("线程已停止。")


def worker(in_queue, out_queue):
    logging.basicConfig(level=logging.DEBUG)
    while True:
        try:
            item = in_queue.get()
            if item is None:
                logging.debug("线程退出")
                break
            # 处理item数组
            excel_sheets = {}
            for excel_name in item:
                sheet_names = excel_utils.get_sheet_names_fast(excel_name)
                excel_sheets[excel_name] = sheet_names
                # logging.debug(f"线程处理工作簿:{excel_name}，页签数：{sheet_names}")
                if excel_sheets.__len__() >= 10:
                    out_queue.put(excel_sheets)
                    # logging.debug(f"线程处理一批工作簿，数量：{excel_sheets.__len__()}")
                    excel_sheets = {}
            if excel_sheets.__len__() > 0:
                out_queue.put(excel_sheets)
                # logging.debug(f"线程处理一批工作簿，数量：{excel_sheets.__len__()}")
        except Exception as e:
            logging.error(f"线程处理异常：{e}")


def get_all_sheet_names():
    use_path = get_config_value("usePath")
    sheet_name = get_path_sheet_name(use_path)
    handlerRead = get_cache_sheet(sheet_name)
    data = handlerRead.get_all_data()
    # 遍历 data 转为字典，key 为文件名，value 为行数据
    sheet_names = []
    for row in data:
        sheets = row["sheets"]
        if sheets:
            for sheet_name in sheets.split('\n'):
                sheet_names.append({"name": row["name"], "sheet_name": sheet_name})
    return sheet_names


# 过滤不包含|的页签名，并转为\n分隔的字符串
def filter_sheet_names(sheet_names):
    # return "\n".join([sheet_name for sheet_name in sheet_names if "|" in sheet_name])
    return "\n".join([sheet_name for sheet_name in sheet_names])


# 判断当前缓存是否存在页签
def exist_sheet(sheet_name):
    names = get_excel_sheet_names(CACHE_EXCEL_NAME)
    return sheet_name in names


if __name__ == '__main__':
    excel_name = "cache.xlsx"
    sheet_name = "缓存文件列表"
    # 计算缓存数据
    handler = ExcelSheetHandler(excel_name, sheet_name)
    handler.clear_data_rows()
    handler.save_workbook()
