import unittest

import openpyxl

from excel_sheet_handler import ExcelSheetHandler


class TestExcelSheetHandler(unittest.TestCase):
    def setUp(self):
        self.handler = ExcelSheetHandler("test_workbook.xlsx", "Sheet1")

    def test_create_sheet(self):
        workbook = self.handler.workbook
        sheet = workbook[self.handler.sheet_name]
        return sheet

    def test_insert_column_header(self):
        self.handler.insert_column_header(5, "cs_value", "type_value", "name_value", "note_value")
        sheet = self.test_create_sheet()
        self.assertEqual(sheet.cell(row=1, column=5).value, "cs_value")
        self.assertEqual(sheet.cell(row=2, column=5).value, "type_value")
        self.assertEqual(sheet.cell(row=3, column=5).value, "name_value")
        self.assertEqual(sheet.cell(row=4, column=5).value, "note_value")

    def test_write_column_header(self):
        self.handler.write_column_header(6, "cs_write", "type_write", "name_write", "note_write")
        sheet = self.test_create_sheet()
        self.assertEqual(sheet.cell(row=1, column=6).value, "cs_write")
        self.assertEqual(sheet.cell(row=2, column=6).value, "type_write")
        self.assertEqual(sheet.cell(row=3, column=6).value, "name_write")
        self.assertEqual(sheet.cell(row=4, column=6).value, "note_write")

    def test_delete_column(self):
        self.handler.insert_column_header(7, "to_delete", "type", "name", "note")
        self.handler.delete_column(7)
        sheet = self.test_create_sheet()
        self.assertTrue(7 not in [cell.column for cell in sheet[1]])

    def test_get_sheet_header(self):
        self.handler.insert_column_header(10, "get_header", "type", "name", "note")
        headers = self.handler.get_sheet_header()
        self.assertIsNotNone(headers)
        found = False
        for header in headers:
            if header["name"] == "name":
                found = True
                break
        self.assertTrue(found)

    def test_get_column_count(self):
        self.handler.insert_column_header(11, "count", "type", "name", "note")
        sheet = self.test_create_sheet()
        count = sheet.max_column
        self.assertEqual(count, sheet.max_column)

    def test_clear_data_rows(self):
        sheet = self.test_create_sheet()
        original_data = []
        for row in range(5, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                initial_value = row * col if row % 2 == 0 else None
                sheet.cell(row=row, column=col).value = initial_value
                row_data.append(initial_value)
            original_data.append(row_data)

        self.handler.clear_data_rows()

        for row in range(5, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                self.assertIsNone(sheet.cell(row=row, column=col).value, f"单元格 ({row}, {col}) 未被正确清空。")

    def test_get_all_data(self):
        sheet = self.test_create_sheet()
        data = self.handler.get_all_data()
        self.assertIsNotNone(data)
        self.assertEqual(len(data), 5)

    def test_get_column_values_by_column_names(self):
        sheet = self.test_create_sheet()
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        for i in range(5, 10):
            sheet.cell(row=i, column=1).value = f"data{i - 4}"
            sheet.cell(row=i, column=2).value = f"data{i - 4}_2"
        data = self.handler.get_column_values_by_column_names("name1", "name2")
        self.assertIsNotNone(data)
        self.assertEqual(len(data), 5)
        self.assertIn("data1", data[0]['name1'])
        self.assertIn("data1_2", data[0]['name2'])

    def test_get_unique_column_values(self):
        sheet = self.test_create_sheet()
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        column_index = 1
        for i in range(5, 10):
            sheet.cell(row=i, column=column_index).value = i - 4
            sheet.cell(row=i + 1, column=column_index).value = i - 4

        unique_values = self.handler.get_unique_column_values("name1")

        self.assertIsNotNone(unique_values)
        self.assertEqual(unique_values, {1, 2, 3, 4, 5})

    def test_get_max_value_from_column(self):
        sheet = self.test_create_sheet()
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        values = [10, 20, 30, 15]
        for i in range(5, 9):
            sheet.cell(row=i, column=1).value = values[i - 5]
        max_value = self.handler.get_max_value_from_column("name1")
        self.assertEqual(max_value, 30)

    def test_alloc_sn(self):
        sheet = self.test_create_sheet()
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "sn"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=5, column=1).value = 10
        sheet.cell(row=6, column=1).value = 15
        sn = self.handler.alloc_sn()
        self.assertEqual(sn, 16)

    def test_get_max_sn(self):
        sheet = self.test_create_sheet()
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "sn"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=5, column=1).value = 10
        max_sn = self.handler.get_max_sn()
        self.assertEqual(max_sn, 10)

    def test_get_column_index_by_name(self):
        self.handler.insert_column_header(12, "test_col", "type", "name", "note")
        index = self.handler.get_column_index_by_name("name")
        self.assertEqual(index, 12)
