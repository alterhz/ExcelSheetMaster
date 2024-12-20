import logging

import openpyxl


class ExcelSheetHandler:
    # 保存已经打开的工作簿
    static_all_workbook = {}

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = None
        self.load_workbook()
        self.worksheet = None

    def load_workbook(self):
        if not self.check_workbook_exists():
            logging.error(f"无法找到指定的工作簿：{self.file_name}")
            return None
        # static_all_workbook 是一个字典，键是文件名，值是工作簿对象
        if self.file_name in ExcelSheetHandler.static_all_workbook:
            self.workbook = ExcelSheetHandler.static_all_workbook[self.file_name]
            logging.debug(f"打开工作簿：{self.file_name}")
        else:
            self.workbook = openpyxl.load_workbook(self.file_name)
            ExcelSheetHandler.static_all_workbook[self.file_name] = self.workbook
            logging.debug(f"加载工作簿：{self.file_name}")

        return self.workbook

    @staticmethod
    def close_all_workbook():
        for workbook_name, workbook in ExcelSheetHandler.static_all_workbook.items():
            workbook.close()
            logging.debug(f"关闭工作簿：{workbook_name}")
        ExcelSheetHandler.static_all_workbook.clear()

    def check_workbook_exists(self):
        try:
            with open(self.file_name, 'rb') as f:
                return True
        except FileNotFoundError:
            return False

    def create_sheet(self, can_create=True):
        if self.workbook is None:
            logging.error(f"无法打开工作簿 {self.file_name}。")
            return None
        try:
            if self.worksheet is not None:
                return self.worksheet
            if self.sheet_name not in self.workbook.sheetnames:
                if can_create:
                    self.workbook.create_sheet(self.sheet_name)
                    logging.info(f"在工作簿 {self.file_name} 中创建工作表 {self.sheet_name}。")
                else:
                    logging.error(f"在工作簿 {self.file_name} 中找不到工作表 {self.sheet_name}。")
                    return None
            logging.debug(f"打开工作簿 {self.file_name} 的工作表 {self.sheet_name}。")
            self.worksheet = self.workbook[self.sheet_name]
            return self.worksheet
        except FileNotFoundError:
            print(f"打开工作簿 {self.file_name} 的工作表 {self.sheet_name} 时出现错误。")
            return None

    def insert_column_header(self, index, cs, type_name, name, note, column_width=0):
        sheet = self.create_sheet()
        if sheet:
            sheet.cell(row=1, column=index).value = cs
            sheet.cell(row=2, column=index).value = type_name
            sheet.cell(row=3, column=index).value = name
            sheet.cell(row=4, column=index).value = note
            # 自动换行
            sheet.cell(row=4, column=index).alignment = openpyxl.styles.Alignment(wrap_text=True)
            if column_width >= 0:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = column_width
            print(f"插入列。index:{index}, name:{name}")

    def write_column_header(self, index, cs, type_name, name, note, column_width=0):
        sheet = self.create_sheet()
        if sheet:
            sheet.cell(row=1, column=index).value = cs
            sheet.cell(row=2, column=index).value = type_name
            sheet.cell(row=3, column=index).value = name
            sheet.cell(row=4, column=index).value = note
            if column_width >= 0:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = column_width
            print(f"写入列。index:{index}, name:{name}")

    def set_header_color(self, color):
        sheet = self.create_sheet()
        if sheet:
            for row in range(1, 5):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(start_color=color,
                                                                                       end_color=color,
                                                                                       fill_type="solid")

    def delete_column(self, index):
        sheet = self.create_sheet()
        if sheet:
            if 1 <= index <= sheet.max_column:
                sheet.delete_cols(index)
            else:
                print("指定的列索引无效")

    def move_column(self, from_index, to_index):
        sheet = self.create_sheet()
        if sheet:
            if 1 <= from_index <= sheet.max_column and 1 <= to_index <= sheet.max_column:
                column_to_move = [row[from_index - 1].value for row in sheet.iter_rows(values_only=True)]
                sheet.delete_cols(from_index)
                sheet.insert_cols(to_index)
                for row_num, value in enumerate(column_to_move, start=1):
                    sheet.cell(row=row_num, column=to_index).value = value
            else:
                print("指定的列索引无效")

    def get_sheet_header(self):
        sheet = self.create_sheet()
        if sheet:
            data_collection = []
            column_count = sheet.max_column
            for i in range(1, column_count + 1):
                column_data = {
                    "columnIndex": i,
                    "exportType": sheet.cell(row=1, column=i).value,
                    "dataType": sheet.cell(row=2, column=i).value,
                    "name": sheet.cell(row=3, column=i).value,
                    "note": sheet.cell(row=4, column=i).value
                }
                data_collection.append(column_data)
            return data_collection
        else:
            return None

    def get_column_count(self):
        sheet = self.create_sheet()
        if sheet:
            return sheet.max_column
        else:
            return None

    def clear_data_rows(self):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            if last_row >= 5:
                for row in range(last_row, 4, -1):
                    sheet.delete_rows(row)
        return None

    def get_first_row_data_by_column_values(self, *find_values):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            result_dict = {}
            for i in range(1, last_row + 1):
                all_columns_match = True
                j = 0
                while j < len(find_values):
                    column_name = find_values[j]
                    value_to_find = find_values[j + 1]
                    column_index = self.get_column_index_by_name(column_name)
                    if column_index == 0:
                        print(f"找不到指定的列名：{column_name}")
                        return {}
                    cell_value = sheet.cell(row=i, column=column_index).value
                    if cell_value != value_to_find:
                        all_columns_match = False
                        break
                    j += 2
                if all_columns_match:
                    column_count = sheet.max_column
                    for column_index in range(1, column_count + 1):
                        key = sheet.cell(row=3, column=column_index).value
                        value = sheet.cell(row=i, column=column_index).value
                        result_dict[key] = value
                    result_dict["row"] = i
                    break
            return result_dict
        else:
            return {}

    def get_last_row_data_by_column_values(self, *find_values):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            result_dict = {}
            found = False
            if len(find_values) == 0:
                found_row = last_row
                found = True
            else:
                for i in range(last_row, 0, -1):
                    all_columns_match = True
                    j = 0
                    while j < len(find_values):
                        column_name = find_values[j]
                        value_to_find = find_values[j + 1]
                        column_index = self.get_column_index_by_name(column_name)
                        if column_index == 0:
                            print(f"找不到指定的列名：{column_name}")
                            return {}
                        cell_value = sheet.cell(row=i, column=column_index).value
                        if cell_value != value_to_find:
                            all_columns_match = False
                            break
                        j += 2
                    if all_columns_match:
                        found_row = i
                        found = True
                        break
            if found:
                column_count = sheet.max_column
                for column_index in range(1, column_count + 1):
                    key = sheet.cell(row=3, column=column_index).value
                    value = sheet.cell(row=found_row, column=column_index).value
                    result_dict[key] = value
                result_dict["row"] = found_row
            return result_dict
        else:
            return {}

    def get_all_data(self):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            result_collection = []
            for i in range(5, last_row + 1):
                result_dict = {}
                column_count = sheet.max_column
                for column_index in range(1, column_count + 1):
                    key = sheet.cell(row=3, column=column_index).value
                    value = sheet.cell(row=i, column=column_index).value
                    result_dict[key] = value
                result_dict["row"] = i
                result_collection.append(result_dict)
            return result_collection
        else:
            return None

    def get_all_rows_data_by_column_values(self, *find_values):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            result_collection = []
            found = False
            if len(find_values) == 0:
                found = True
            else:
                result = find_values
                for i in range(5, last_row + 1):
                    all_columns_match = True
                    j = 0
                    while j < len(result):
                        column_name = result[j]
                        value_to_find = result[j + 1]
                        column_index = self.get_column_index_by_name(column_name)
                        if column_index == 0:
                            print(f"找不到指定的列名：{column_name}")
                            return []
                        cell_value = sheet.cell(row=i, column=column_index).value
                        if str(cell_value) != str(value_to_find):
                            all_columns_match = False
                            break
                        j += 2
                    if all_columns_match:
                        found = True
                        result_dict = {}
                        column_count = sheet.max_column
                        for column_index in range(1, column_count + 1):
                            key = sheet.cell(row=3, column=column_index).value
                            value = sheet.cell(row=i, column=column_index).value
                            result_dict[key] = value
                        result_dict["row"] = i
                        result_collection.append(result_dict)
            if found:
                return result_collection
            else:
                return None
        else:
            return None

    def get_column_values_by_column_names(self, *column_names):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            temp_col = []
            for i in range(5, last_row + 1):
                temp_dict = {"row": i}
                for column_name in column_names:
                    column_index = self.get_column_index_by_name(column_name)
                    if column_index == 0:
                        print(f"找不到指定的列名：{column_name}")
                        return None
                    temp_dict[column_name] = sheet.cell(row=i, column=column_index).value
                temp_col.append(temp_dict)
            return temp_col
        else:
            return None

    def get_unique_column_values(self, column_name):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            unique_values = set()
            column_index = self.get_column_index_by_name(column_name)
            if column_index == 0:
                print(f"找不到指定的列名：{column_name}")
                return None
            for i in range(5, last_row + 1):
                temp_value = sheet.cell(row=i, column=column_index).value
                if temp_value is not None:
                    unique_values.add(temp_value)
            return unique_values
        else:
            return None

    def is_in_collection(self, value, col):
        for item in col:
            if item == value:
                return True
        return False

    def get_max_value_from_column(self, column_name):
        sheet = self.create_sheet()
        if sheet:
            column_index = self.get_column_index_by_name(column_name)
            if column_index == 0:
                print(f"找不到指定的列名：{column_name}")
                return None
            max_value = None
            for i in range(5, sheet.max_row - 1):
                value = sheet.cell(row=i, column=column_index).value
                if value is not None:
                    try:
                        int_value = int(value)
                        if max_value is None or int_value > max_value:
                            max_value = int_value
                    except ValueError:
                        print(f"单元格的值 '{value}' 不是数值类型，无法处理。")
            if max_value is not None:
                print(f"在列 '{column_name}' 中找到最大值为 {max_value}。")
                return max_value
            else:
                print(f"在列 '{column_name}' 中未找到任何数值，返回 None。")
                return None
        else:
            print("无法获取工作表，返回 None。")
            return None

    def alloc_sn(self):
        max_sn = self.get_max_value_from_column("sn")
        return max_sn + 1

    def get_max_sn(self):
        max_sn = self.get_max_value_from_column("sn")
        return max_sn

    def write_row_data(self, row_number, dict_data, insert=False):
        sheet = self.create_sheet()
        if sheet:
            if row_number == -1:
                row_number = sheet.max_row + 1
            elif row_number < 5:
                print("插入行号不能小于 5。")
                return -1
            if insert:
                sheet.insert_rows(row_number)
            for key, value in dict_data.items():
                column_name = key
                column_index = self.get_column_index_by_name(column_name)
                if column_index != 0:
                    data_type_name = self.get_data_type_name(column_name)
                    target_cell = sheet.cell(row=row_number, column=column_index)
                    if "string" in data_type_name.lower() or "[]" in data_type_name:
                        target_cell.number_format = "@"
                    target_cell.value = value
                    # 设置自动换行
                    target_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            return row_number
        else:
            return -1

    def insert_row_and_write_param_data(self, row_number, *column_value_pairs):
        sheet = self.create_sheet()
        if sheet:
            if row_number == -1:
                row_number = sheet.max_row + 1
            sheet.insert_rows(row_number)
            i = 0
            while i < len(column_value_pairs):
                column_name = column_value_pairs[i]
                value = column_value_pairs[i + 1]
                column_index = self.get_column_index_by_name(column_name)
                if column_index != 0:
                    sheet.cell(row=row_number, column=column_index).value = value
                i += 2

    def delete_range_rows(self, start_row, end_row=None):
        sheet = self.create_sheet()
        if sheet:
            if end_row is None:
                sheet.delete_rows(start_row)
            else:
                for row in range(end_row, start_row - 1, -1):
                    sheet.delete_rows(row)
        return None

    def delete_rows(self, row_numbers: list):
        """
        删除行,按行号从大到小删除
        """
        sheet = self.create_sheet()
        if sheet:
            sorted_rows = sorted(row_numbers, reverse=True)
            for row_number in sorted_rows:
                if 1 <= row_number <= sheet.max_row:
                    sheet.delete_rows(row_number)
        return None

    def get_column_index_by_name(self, column_name):
        sheet = self.create_sheet()
        if sheet:
            last_column = sheet.max_column
            for i in range(1, last_column + 1):
                if sheet.cell(row=3, column=i).value == column_name:
                    return i
            return 0
        else:
            return 0

    def get_data_type_name(self, column_name):
        sheet = self.create_sheet()
        if sheet:
            last_column = sheet.max_column
            for i in range(1, last_column + 1):
                if sheet.cell(row=3, column=i).value == column_name:
                    return sheet.cell(row=2, column=i).value
            return None
        else:
            return None

    def move_to_row(self, row_number):
        sheet = self.create_sheet()
        if sheet:
            last_row = sheet.max_row
            if row_number < 1:
                row_number = 1
            elif row_number > last_row:
                row_number = last_row
            sheet.cell(row=row_number, column=1).select()

    def set_row_color(self, row_number, color):
        sheet = self.create_sheet()
        if sheet:
            for col_num in range(1, sheet.max_column + 1):
                sheet.cell(row=row_number, column=col_num).fill = openpyxl.styles.PatternFill(start_color=color,
                                                                                              end_color=color,
                                                                                              fill_type="solid")

    def move_to_row_and_set_color(self, row_number, color):
        self.move_to_row(row_number)
        self.set_row_color(row_number, color)

    def get_max_row_number(self):
        sheet = self.create_sheet()
        if sheet:
            return sheet.max_row
        else:
            return None

    def save_workbook(self):
        if self.workbook:
            self.workbook.save(self.file_name)
            logging.debug(f"保存工作簿：{self.file_name}, 页签：{self.sheet_name}")
        else:
            print(f"无法保存工作簿，未找到指定的工作簿：{self.file_name}")

    def remove_sheet(self):
        if self.workbook:
            self.workbook.remove(self.worksheet)
            self.save_workbook()
            logging.debug(f"删除工作簿：{self.file_name}, 页签：{self.sheet_name}")
        else:
            print(f"无法删除工作簿，未找到指定的工作簿：{self.file_name}")


if __name__ == '__main__':
    handler = ExcelSheetHandler("your_workbook_name.xlsx", "your_sheet_name")
    # 调用各种方法进行测试
