import unittest

import openpyxl

from ExcelSheetHandler import ExcelSheetHandler


class TestExcelSheetHandler(unittest.TestCase):
    def setUp(self):
        self.handler = ExcelSheetHandler("test_workbook.xlsx", "Sheet1")

    def test_create_sheet(self):
        workbook = self.handler.create_sheet()
        self.assertIsNotNone(workbook)

    def test_insert_column_header(self):
        self.handler.insert_column_header(5, "cs_value", "type_value", "name_value", "note_value")
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=1, column=5).value, "cs_value")
        self.assertEqual(sheet.cell(row=2, column=5).value, "type_value")
        self.assertEqual(sheet.cell(row=3, column=5).value, "name_value")
        self.assertEqual(sheet.cell(row=4, column=5).value, "note_value")

    def test_write_column_header(self):
        self.handler.write_column_header(6, "cs_write", "type_write", "name_write", "note_write")
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=1, column=6).value, "cs_write")
        self.assertEqual(sheet.cell(row=2, column=6).value, "type_write")
        self.assertEqual(sheet.cell(row=3, column=6).value, "name_write")
        self.assertEqual(sheet.cell(row=4, column=6).value, "note_write")

    def test_set_header_color(self):
        self.handler.set_header_color("FF0000")
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        for row in range(1, 5):
            for col in range(1, sheet.max_column + 1):
                self.assertEqual(sheet.cell(row=row, column=col).fill.start_color.rgb, "FF0000")

    def test_delete_column(self):
        self.handler.insert_column_header(7, "to_delete", "type", "name", "note")
        self.handler.delete_column(7)
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertTrue(7 not in [cell.column for cell in sheet[1]])

    def test_move_column(self):
        self.handler.insert_column_header(8, "original", "type", "name", "note")
        self.handler.insert_column_header(9, "target", "type", "name", "note")
        self.handler.move_column(8, 9)
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=1, column=9).value, "original")

    def test_get_sheet_header(self):
        self.handler.insert_column_header(10, "get_header", "type", "name", "note")
        headers = self.handler.get_sheet_header()
        self.assertIsNotNone(headers)
        found = False
        for header in headers:
            if header["name"] == "name":
                found = True
                break
        self.assertTrue(found)

    def test_get_column_count(self):
        self.handler.insert_column_header(11, "count", "type", "name", "note")
        count = self.handler.get_column_count()
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(count, sheet.max_column)

    def test_clear_data_rows(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for i in range(1, 10):
            sheet.cell(row=i, column=1).value = i
        workbook.save("test_workbook.xlsx")
        self.handler.clear_data_rows()
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        for row in range(5, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                self.assertIsNone(sheet.cell(row=row, column=col).value)

    def test_get_first_row_data_by_column_values(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        sheet.cell(row=1, column=3).value = "search_col"
        sheet.cell(row=2, column=3).value = "type3"
        sheet.cell(row=3, column=3).value = "search_name"
        sheet.cell(row=4, column=3).value = "note3"
        workbook.save("test_workbook.xlsx")
        data = self.handler.get_first_row_data_by_column_values("search_col", "search_name")
        self.assertIsNotNone(data)
        self.assertEqual(data["name1"], "name1")
        self.assertEqual(data["name2"], "name2")
        self.assertEqual(data["search_name"], "search_name")

    def test_get_last_row_data_by_column_values(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        sheet.cell(row=1, column=3).value = "search_col"
        sheet.cell(row=2, column=3).value = "type3"
        sheet.cell(row=3, column=3).value = "search_name"
        sheet.cell(row=4, column=3).value = "note3"
        sheet.cell(row=1, column=4).value = "col4"
        sheet.cell(row=2, column=4).value = "type4"
        sheet.cell(row=3, column=4).value = "name4"
        sheet.cell(row=4, column=4).value = "note4"
        workbook.save("test_workbook.xlsx")
        data = self.handler.get_last_row_data_by_column_values("search_col", "search_name")
        self.assertIsNotNone(data)
        self.assertEqual(data["name1"], "name1")
        self.assertEqual(data["name2"], "name2")
        self.assertEqual(data["name4"], "name4")
        self.assertEqual(data["search_name"], "search_name")

    def test_get_all_data(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        for i in range(5, 10):
            sheet.cell(row=i, column=1).value = f"data{i - 4}"
            sheet.cell(row=i, column=2).value = f"data{i - 4}_2"
        workbook.save("test_workbook.xlsx")
        data = self.handler.get_all_data()
        self.assertIsNotNone(data)
        self.assertEqual(len(data), 5)

    def test_get_all_rows_data_by_column_values(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        for i in range(5, 10):
            sheet.cell(row=i, column=1).value = f"data{i - 4}"
            sheet.cell(row=i, column=2).value = f"data{i - 4}_2"
            if i == 7:
                sheet.cell(row=i, column=1).value = "search_value"
                sheet.cell(row=i, column=2).value = "search_value_2"
        workbook.save("test_workbook.xlsx")
        data = self.handler.get_all_rows_data_by_column_values("col1", "search_value")
        self.assertIsNotNone(data)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["name1"], "name1")
        self.assertEqual(data[0]["name2"], "name2")

    def test_get_column_values_by_column_names(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=1, column=2).value = "col2"
        sheet.cell(row=2, column=2).value = "type2"
        sheet.cell(row=3, column=2).value = "name2"
        sheet.cell(row=4, column=2).value = "note2"
        for i in range(5, 10):
            sheet.cell(row=i, column=1).value = f"data{i - 4}"
            sheet.cell(row=i, column=2).value = f"data{i - 4}_2"
        workbook.save("test_workbook.xlsx")
        data = self.handler.get_column_values_by_column_names("col1", "col2")
        self.assertIsNotNone(data)
        self.assertEqual(len(data), 5)
        self.assertIn("col1", data[0])
        self.assertIn("col2", data[0])

    def test_get_unique_column_values(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        values = ["value1", "value2", "value1", "value3"]
        for i in range(5, 9):
            sheet.cell(row=i, column=1).value = values[i - 5]
        workbook.save("test_workbook.xlsx")
        unique_values = self.handler.get_unique_column_values("col1")
        self.assertIsNotNone(unique_values)
        self.assertEqual(len(unique_values), 3)
        self.assertIn("value1", unique_values)
        self.assertIn("value2", unique_values)
        self.assertIn("value3", unique_values)

    def test_get_max_value_from_column(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "name1"
        sheet.cell(row=4, column=1).value = "note1"
        values = [10, 20, 30, 15]
        for i in range(5, 9):
            sheet.cell(row=i, column=1).value = values[i - 5]
        workbook.save("test_workbook.xlsx")
        max_value = self.handler.get_max_value_from_column("col1")
        self.assertEqual(max_value, 30)

    def test_alloc_sn(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "sn"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=5, column=1).value = 10
        sheet.cell(row=6, column=1).value = 15
        workbook.save("test_workbook.xlsx")
        sn = self.handler.alloc_sn()
        self.assertEqual(sn, 16)

    def test_get_max_sn(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.cell(row=1, column=1).value = "col1"
        sheet.cell(row=2, column=1).value = "type1"
        sheet.cell(row=3, column=1).value = "sn"
        sheet.cell(row=4, column=1).value = "note1"
        sheet.cell(row=5, column=1).value = 10
        sheet.cell(row=6, column=1).value = 15
        workbook.save("test_workbook.xlsx")
        max_sn = self.handler.get_max_sn()
        self.assertEqual(max_sn, 15)

    def test_insert_row_and_write_dict_data(self):
        data = {"col1": "new_value", "col2": "new_value_2"}
        row_number = self.handler.insert_row_and_write_dict_data(-1, data)
        self.assertGreater(row_number, 0)
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=row_number, column=1).value, "new_value")
        self.assertEqual(sheet.cell(row=row_number, column=2).value, "new_value_2")

    def test_insert_row_and_write_param_data(self):
        self.handler.insert_row_and_write_param_data(-1, "col1", "new_value", "col2", "new_value_2")
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        last_row = sheet.max_row
        self.assertEqual(sheet.cell(row=last_row, column=1).value, "new_value")
        self.assertEqual(sheet.cell(row=last_row, column=2).value, "new_value_2")

    def test_get_column_index_by_name(self):
        self.handler.insert_column_header(12, "test_col", "type", "name", "note")
        index = self.handler.get_column_index_by_name("test_col")
        self.assertEqual(index, 12)

    def test_move_to_row(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for i in range(1, 10):
            sheet.cell(row=i, column=1).value = i
        workbook.save("test_workbook.xlsx")
        self.handler.move_to_row(7)
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=7, column=1).selected, True)

    def test_set_row_color(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for i in range(1, 10):
            sheet.cell(row=i, column=1).value = i
        workbook.save("test_workbook.xlsx")
        self.handler.set_row_color(5, "FF0000")
        workbook = self.handler.create_sheet()
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=5, column=1).fill.start_color.rgb, "FF0000")

